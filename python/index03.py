import pandas as pd
import requests as ree
from openpyxl import load_workbook
from openpyxl import Workbook
import re

excel_file = r'C:/Users/qetqe/Documents/github/gaboza/python/University.xlsx'

df_from_xl = pd.read_excel(excel_file, engine="openpyxl")
df_from_xl.columns = df_from_xl.loc[4].tolist()
df_from_xl = df_from_xl.drop(index=list(range(0, 5)))

url="http://api.vworld.kr/req/address?"
# 디버깅 - Params에 URL GET 오류
params='service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type='ROAD'
number_type='PARCEL'
address='&address='
key='&key='
primary_key='FAC47DA3-B363-397D-BD3D-4573B009CAC5'

def request_geo(road):
    page = ree.get(url+params+road_type+address+road+key+primary_key)
    json_data=page.json()
    
    if json_data['response']['status'] == "OK":
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        x = 0
        y = 0
        return x, y

try:
    wb = load_workbook(r"C:/Users/qetqe/Documents/github/gaboza/python/학교주소좌표.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_xl['학교명'].to_list()
address_list = df_from_xl['주소'].to_list()

for num,value in enumerate(address_list):
    addr = re.sub(r'\([^)]*\)', '', value)
    x,y = request_geo(addr)
    sheet.append([university_list[num],addr,x,y])

wb.save(r"C:/Users/qetqe/Documents/github/gaboza/python/학교주소좌표.xlsx")